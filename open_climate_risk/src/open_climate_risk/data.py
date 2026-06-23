"""Hazard and vulnerability data loading."""
from pathlib import Path

import numpy as np
import pandas as pd
import openpyxl
import rasterio
from scipy.interpolate import interp1d

from .config import AQUEDUCT_S3, DATA_DIR, JRC_EXCEL, RETURN_PERIODS


# ── Aqueduct helpers ───────────────────────────────────────────────────────────

def aqueduct_filename(
    return_period: int,
    scenario: str = 'historical',
    model: str = '000000000WATCH',
    year: int = 1980,
) -> str:
    """Build Aqueduct v2 riverine flood filename."""
    return f'inunriver_{scenario}_{model}_{year}_rp0{return_period:04d}.tif'


def aqueduct_source(
    return_period: int,
    scenario: str = 'historical',
    model: str = '000000000WATCH',
    year: int = 1980,
    data_dir: Path = None,
) -> str:
    """Return local path if the file exists, else the S3 URL (COG streaming)."""
    data_dir = data_dir or DATA_DIR
    fname = aqueduct_filename(return_period, scenario, model, year)
    local = Path(data_dir) / fname
    if local.exists():
        return str(local)
    return f'{AQUEDUCT_S3}/{fname}'


def check_availability(
    scenarios: list,
    return_periods: np.ndarray = None,
    data_dir: Path = None,
) -> None:
    """Print a file-availability table for all scenarios × models × RPs."""
    return_periods = return_periods if return_periods is not None else RETURN_PERIODS
    data_dir = data_dir or DATA_DIR
    for sc in scenarios:
        print(f"  [{sc['label']}]")
        for model in sc['models']:
            n_ok = sum(
                (Path(data_dir) / aqueduct_filename(int(rp), sc['scenario'], model, sc['year'])).exists()
                for rp in return_periods
            )
            marker = '✓' if n_ok == len(return_periods) else ('~' if n_ok > 0 else '✗')
            print(f'    {marker} {model}  ({n_ok}/{len(return_periods)} fichiers présents)')
        print()


# ── Coastal flood helpers ──────────────────────────────────────────────────────

def coastal_filename(
    return_period: int,
    scenario: str = 'historical',
    subsidence: str = 'nosub',
    year='hist',
    projection: str = '0',
) -> str:
    """Build Aqueduct v2 coastal flood filename.

    Projection: '0' = 95th percentile (default), '0_perc_05', '0_perc_50'.
    """
    return f'inuncoast_{scenario}_{subsidence}_{year}_rp{return_period:04d}_{projection}.tif'


def coastal_source(
    return_period: int,
    scenario: str = 'historical',
    subsidence: str = 'nosub',
    year='hist',
    projection: str = '0',
    data_dir: Path = None,
) -> str:
    """Return local path if exists, else S3 URL for coastal flood raster."""
    data_dir = data_dir or DATA_DIR
    fname = coastal_filename(return_period, scenario, subsidence, year, projection)
    local = Path(data_dir) / fname
    if local.exists():
        return str(local)
    return f'{AQUEDUCT_S3}/{fname}'


def check_coastal_availability(
    scenarios: list,
    return_periods: np.ndarray = None,
    data_dir: Path = None,
) -> None:
    """Print file-availability for coastal scenarios."""
    return_periods = return_periods if return_periods is not None else RETURN_PERIODS
    data_dir = data_dir or DATA_DIR
    for sc in scenarios:
        print(f"  [{sc['label']}]")
        n_ok = sum(
            (Path(data_dir) / coastal_filename(
                int(rp), sc['scenario'], sc['subsidence'], sc['year'], sc['projection']
            )).exists()
            for rp in return_periods
        )
        marker = '\u2713' if n_ok == len(return_periods) else ('~' if n_ok > 0 else '\u2717')
        print(f'    {marker} {sc["subsidence"]}/{sc["projection"]}  '
              f'({n_ok}/{len(return_periods)} fichiers)')
    print()


def extract_coastal_depths(
    lon: float,
    lat: float,
    return_periods: np.ndarray = None,
    data_dir: Path = None,
    scenario: str = 'historical',
    subsidence: str = 'nosub',
    year='hist',
    projection: str = '0',
) -> np.ndarray:
    """Extract coastal flood depth [m] for each return period."""
    return_periods = return_periods if return_periods is not None else RETURN_PERIODS
    data_dir = data_dir or DATA_DIR
    depths = np.full(len(return_periods), np.nan)
    n_remote = 0

    for i, rp in enumerate(return_periods):
        source = coastal_source(int(rp), scenario, subsidence, year, projection, data_dir)
        if source.startswith('http'):
            n_remote += 1
        try:
            depths[i] = sample_raster_at_point(source, lon, lat)
        except Exception:
            depths[i] = np.nan

    if n_remote > 0:
        print(f'  \u2601  {n_remote}/{len(return_periods)} COG stream\u00e9s ({scenario}/{subsidence})')
    if np.all(np.isnan(depths)):
        print('\u26a0  Aucune donn\u00e9e coastal disponible.')
        return np.zeros(len(return_periods))
    return np.where(np.isnan(depths), 0.0, depths)


def extract_coastal_df(
    asset: dict,
    return_periods: np.ndarray = None,
    data_dir: Path = None,
    subsidence: str = 'nosub',
    projection: str = '0',
) -> 'pd.DataFrame':
    """Extract coastal hazard depths for an asset → tidy DataFrame."""
    return_periods = return_periods if return_periods is not None else RETURN_PERIODS
    depths = extract_coastal_depths(
        asset['lon'], asset['lat'], return_periods, data_dir,
        subsidence=subsidence, projection=projection,
    )
    return pd.DataFrame({
        'return_period_yr':   return_periods,
        'exceedance_prob_yr': 1.0 / return_periods,
        'flood_depth_m':      depths,
    })


# ── Raster sampling ────────────────────────────────────────────────────────────

def sample_raster_at_point(source: str, lon: float, lat: float) -> float:
    """Sample a GeoTIFF (local or S3 URL) at a lon/lat point. Returns 0.0 if nodata."""
    with rasterio.open(source) as src:
        row, col = src.index(lon, lat)
        data = src.read(1, window=rasterio.windows.Window(col, row, 1, 1))
        nodata = src.nodata
        value = float(data[0, 0])
        if nodata is not None and value == nodata:
            return 0.0
        return max(0.0, value)


def _synthetic_hazard_curve(return_periods: np.ndarray) -> np.ndarray:
    anchor_T = np.array([2, 10, 100, 1000], dtype=float)
    anchor_d = np.array([0.0, 0.3, 1.2, 2.8])
    f = interp1d(np.log(anchor_T), anchor_d, kind='linear',
                 bounds_error=False, fill_value=(anchor_d[0], anchor_d[-1]))
    return np.maximum(0.0, f(np.log(return_periods)))


def extract_hazard_depths(
    lon: float,
    lat: float,
    return_periods: np.ndarray = None,
    data_dir: Path = None,
    scenario: str = 'historical',
    model: str = '000000000WATCH',
    year: int = 1980,
) -> np.ndarray:
    """Extract flood depth [m] for each return period. Streams from S3 if not local."""
    return_periods = return_periods if return_periods is not None else RETURN_PERIODS
    data_dir = data_dir or DATA_DIR
    depths = np.full(len(return_periods), np.nan)
    n_remote = 0

    for i, rp in enumerate(return_periods):
        source = aqueduct_source(int(rp), scenario, model, year, data_dir)
        if source.startswith('http'):
            n_remote += 1
        try:
            depths[i] = sample_raster_at_point(source, lon, lat)
        except Exception:
            depths[i] = np.nan

    if n_remote > 0:
        print(f'  ☁  {n_remote}/{len(return_periods)} fichiers streamés depuis S3 ({scenario}/{model})')
    if np.all(np.isnan(depths)):
        print('⚠  Aucune donnée — courbe synthétique utilisée.')
        return _synthetic_hazard_curve(return_periods)

    return depths


def extract_hazard_df(
    asset: dict,
    return_periods: np.ndarray = None,
    data_dir: Path = None,
) -> pd.DataFrame:
    """Extract hazard depths for an asset and return a tidy DataFrame."""
    return_periods = return_periods if return_periods is not None else RETURN_PERIODS
    depths = extract_hazard_depths(asset['lon'], asset['lat'], return_periods, data_dir)
    return pd.DataFrame({
        'return_period_yr':   return_periods,
        'exceedance_prob_yr': 1.0 / return_periods,
        'flood_depth_m':      depths,
    })


# ── JRC vulnerability curves ───────────────────────────────────────────────────

def load_jrc_curves(excel_path: Path = None, region: str = 'EUROPE') -> dict:
    """Load JRC Huizinga 2017 relative damage curves from the official Excel."""
    excel_path = excel_path or JRC_EXCEL

    ASSET_MAP = {
        'Residential buildings': 'residential',
        'Commercial buildings':  'commercial',
        'Industrial buildings':  'industrial',
        'Agriculture':           'agriculture',
        'Transport':             'transport',
    }
    COLORS = {
        'residential': '#e63946', 'commercial': '#2196F3',
        'industrial':  '#FF9800', 'agriculture': '#4CAF50',
        'transport':   '#9C27B0',
    }

    if not Path(str(excel_path)).exists():
        return _jrc_fallback_curves()

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb['Damage functions']
    header = [str(c.value).replace('\n', ' ').strip() if c.value else ''
              for c in list(ws.iter_rows())[2]]
    region_col = next((i for i, h in enumerate(header) if region.upper() in h.upper()), None)
    if region_col is None:
        raise ValueError(f"Region '{region}' not found in header.")

    curves = {}
    current_type, depths, fracs, done = None, [], [], False

    for row in ws.iter_rows(values_only=True):
        label = row[0]
        if label and str(label) in ASSET_MAP:
            if current_type and depths:
                curves[current_type] = {
                    'depth_m': np.array(depths), 'damage_fraction': np.array(fracs),
                    'color': COLORS[current_type], 'label': current_type.title(),
                }
            current_type, depths, fracs, done = ASSET_MAP[str(label)], [], [], False
            continue
        if not current_type or done:
            continue
        depth = row[1]
        frac  = row[region_col] if region_col < len(row) else None
        if not isinstance(depth, (int, float)) or not isinstance(frac, (int, float)):
            continue
        if depths and float(depth) < depths[-1]:
            done = True
            continue
        depths.append(float(depth))
        fracs.append(float(frac))

    if current_type and depths:
        curves[current_type] = {
            'depth_m': np.array(depths), 'damage_fraction': np.array(fracs),
            'color': COLORS[current_type], 'label': current_type.title(),
        }
    return curves


def load_jrc_max_damage(excel_path: Path = None, country_iso3: str = 'FRA') -> dict:
    """Load JRC max damage values [€/m², 2010] for a given country (ISO3 code)."""
    excel_path = excel_path or JRC_EXCEL
    if not Path(str(excel_path)).exists():
        return {'residential': 776, 'commercial': 1076, 'industrial': 873,
                'agriculture': 0.11, 'transport': 710}

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    iso_lookup = {str(r[1]).strip().upper(): str(r[0]).strip()
                  for r in wb['ISO_Table'].iter_rows(values_only=True)
                  if r[0] and r[1]}
    target = iso_lookup.get(country_iso3.upper())

    result = {}
    for atype, sheet in [('residential', 'MaxDamage-Residential'),
                          ('commercial',  'MaxDamage-Commercial'),
                          ('industrial',  'MaxDamage-Industrial')]:
        for row in wb[sheet].iter_rows(values_only=True):
            if row[0] and target and target.lower() in str(row[0]).lower():
                if isinstance(row[3], (int, float)):
                    result[atype] = float(row[3])
                break

    for row in wb['MaxDamage-Agriculture'].iter_rows(values_only=True):
        if row[0] and target and target.lower() in str(row[0]).lower():
            if isinstance(row[1], (int, float)):
                result['agriculture'] = float(row[1]) / 10_000
            break

    EU_TRANSPORT_EUR_M2 = 751
    EU_GDP_USD = 43_097
    for row in wb['MaxDamage-Transport'].iter_rows(values_only=True):
        if row[0] and target and target.lower() in str(row[0]).lower():
            if isinstance(row[1], (int, float)):
                result['transport'] = EU_TRANSPORT_EUR_M2 * (float(row[1]) / EU_GDP_USD)
            break

    defaults = {'residential': 776, 'commercial': 1076, 'industrial': 873,
                'agriculture': 0.11, 'transport': 710}
    for k, v in defaults.items():
        result.setdefault(k, v)
    return result


# ── Portfolio loading ───────────────────────────────────────────────────────────

def load_portfolio(csv_path) -> list:
    """Load a portfolio of assets from CSV. Returns list of asset dicts.

    Required CSV columns: name, lon, lat, asset_type, value_eur, floor_area_m2
    Optional: id, country, city
    """
    df = pd.read_csv(csv_path)
    required = {'name', 'lon', 'lat', 'asset_type', 'value_eur', 'floor_area_m2'}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f'Missing columns in CSV: {missing}')
    assets = df.to_dict('records')
    print(f'Portfolio chargé : {len(assets)} actifs, '
          f'valeur totale {sum(a["value_eur"] for a in assets):,.0f} EUR')
    return assets


def batch_sample_raster(
    source: str,
    lons: list,
    lats: list,
) -> np.ndarray:
    """Sample a single raster at multiple lon/lat points (opens file once)."""
    n = len(lons)
    values = np.zeros(n)
    with rasterio.open(source) as src:
        nodata = src.nodata
        for i in range(n):
            try:
                row, col = src.index(lons[i], lats[i])
                data = src.read(1, window=rasterio.windows.Window(col, row, 1, 1))
                val = float(data[0, 0])
                if nodata is not None and val == nodata:
                    val = 0.0
                values[i] = max(0.0, val)
            except Exception:
                values[i] = 0.0
    return values


def _jrc_fallback_curves() -> dict:
    DEPTH_M = np.array([0.0, 0.5, 1.0, 1.5, 2.0, 3.0, 4.0, 5.0, 6.0])
    return {
        'residential': {'depth_m': DEPTH_M, 'damage_fraction': np.array([0.00, 0.25, 0.40, 0.50, 0.60, 0.75, 0.85, 0.95, 1.00]), 'color': '#e63946', 'label': 'Residential'},
        'commercial':  {'depth_m': DEPTH_M, 'damage_fraction': np.array([0.00, 0.15, 0.30, 0.45, 0.55, 0.75, 0.90, 1.00, 1.00]), 'color': '#2196F3', 'label': 'Commercial'},
        'industrial':  {'depth_m': DEPTH_M, 'damage_fraction': np.array([0.00, 0.15, 0.27, 0.40, 0.52, 0.70, 0.85, 1.00, 1.00]), 'color': '#FF9800', 'label': 'Industrial'},
        'agriculture': {'depth_m': DEPTH_M, 'damage_fraction': np.array([0.00, 0.30, 0.55, 0.65, 0.75, 0.85, 0.95, 1.00, 1.00]), 'color': '#4CAF50', 'label': 'Agriculture'},
        'transport':   {'depth_m': DEPTH_M, 'damage_fraction': np.array([0.00, 0.32, 0.54, 0.70, 0.83, 1.00, 1.00, 1.00, 1.00]), 'color': '#9C27B0', 'label': 'Transport'},
    }
